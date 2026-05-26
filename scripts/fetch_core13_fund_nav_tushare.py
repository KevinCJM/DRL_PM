from __future__ import annotations

import argparse
import json
import os
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

import pandas as pd
import yaml


ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = ROOT / "data" / "reports"
PROCESSED_DIR = ROOT / "data" / "processed"

DEFAULT_EXCEL_OUTPUT = REPORTS_DIR / "core13_etf_lof_fund_nav_tushare.xlsx"
DEFAULT_LONG_CSV_OUTPUT = REPORTS_DIR / "core13_etf_lof_fund_nav_tushare_long.csv"
DEFAULT_SUMMARY_CSV_OUTPUT = REPORTS_DIR / "core13_etf_lof_fund_nav_tushare_summary.csv"
DEFAULT_LONG_PARQUET_OUTPUT = PROCESSED_DIR / "core13_etf_lof_fund_nav_tushare_long.parquet"
DEFAULT_WIDE_ADJ_NAV_PARQUET_OUTPUT = PROCESSED_DIR / "core13_wide_adj_nav_tushare.parquet"
DEFAULT_MANIFEST_OUTPUT = REPORTS_DIR / "core13_etf_lof_fund_nav_tushare_manifest.json"

FUND_NAV_FIELDS = [
    "ts_code",
    "ann_date",
    "nav_date",
    "unit_nav",
    "accum_nav",
    "accum_div",
    "net_asset",
    "total_netasset",
    "adj_nav",
]


@dataclass(frozen=True)
class FundAsset:
    ts_code: str
    fund_name: str
    instrument_type: str
    asset_bucket: str


CORE13_ASSETS: tuple[FundAsset, ...] = (
    FundAsset("510300.SH", "沪深300ETF华泰柏瑞", "ETF", "A股宽基/风格"),
    FundAsset("510500.SH", "中证500ETF南方", "ETF", "A股宽基/风格"),
    FundAsset("510050.SH", "上证50ETF华夏", "ETF", "A股宽基/风格"),
    FundAsset("159915.SZ", "创业板ETF易方达", "ETF", "A股宽基/风格"),
    FundAsset("510880.SH", "红利ETF华泰柏瑞", "ETF", "A股宽基/风格"),
    FundAsset("159920.SZ", "恒生ETF华夏", "ETF", "港股/海外权益"),
    FundAsset("513100.SH", "纳指ETF国泰", "ETF", "港股/海外权益"),
    FundAsset("518880.SH", "黄金ETF华安", "ETF", "黄金/贵金属"),
    FundAsset("511010.SH", "国债ETF国泰", "ETF", "债券/固收"),
    FundAsset("511880.SH", "银华日利ETF", "ETF", "现金/货币"),
    FundAsset("513500.SH", "标普500ETF博时", "ETF", "港股/海外权益"),
    FundAsset("160216.SZ", "国泰大宗商品(QDII-LOF)A", "LOF", "商品/能源"),
    FundAsset("160416.SZ", "华安标普全球石油指数(LOF)A", "LOF", "商品/能源"),
)


def load_universe(path: Path) -> tuple[FundAsset, ...]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    records = payload.get("assets")
    if not isinstance(records, list) or not records:
        raise ValueError("CORE13_UNIVERSE_INVALID: assets must be a non-empty list")
    assets: list[FundAsset] = []
    seen: set[str] = set()
    for index, record in enumerate(records):
        if not isinstance(record, dict):
            raise ValueError(f"CORE13_UNIVERSE_INVALID: assets[{index}] must be a mapping")
        ts_code = str(record.get("ts_code", "")).strip()
        if not ts_code:
            raise ValueError(f"CORE13_UNIVERSE_INVALID: assets[{index}].ts_code")
        if ts_code in seen:
            raise ValueError(f"CORE13_UNIVERSE_DUPLICATE: {ts_code}")
        seen.add(ts_code)
        assets.append(
            FundAsset(
                ts_code=ts_code,
                fund_name=str(record.get("name") or ts_code),
                instrument_type=str(record.get("asset_type") or record.get("type") or ""),
                asset_bucket=str(record.get("asset_bucket") or record.get("pool") or ""),
            )
        )
    return tuple(assets)


def resolve_token(env_names: Sequence[str]) -> tuple[str, str]:
    for name in env_names:
        value = os.environ.get(name)
        if value:
            return value, name
    joined = ", ".join(env_names)
    raise RuntimeError(f"TUSHARE_TOKEN_NOT_FOUND: set one of [{joined}] in the runtime environment")


def get_tushare_client(token: str) -> Any:
    import tushare as ts

    ts.set_token(token)
    return ts.pro_api()


def fetch_asset_nav(
    pro: Any,
    asset: FundAsset,
    *,
    market: str,
    start_date: str,
    end_date: str,
    retries: int,
    retry_sleep: float,
) -> pd.DataFrame:
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            frame = pro.fund_nav(
                ts_code=asset.ts_code,
                market=market,
                start_date=start_date,
                end_date=end_date,
                fields=",".join(FUND_NAV_FIELDS),
            )
            if not isinstance(frame, pd.DataFrame) or frame.empty:
                raise RuntimeError(f"EMPTY_FUND_NAV {asset.ts_code}")
            return normalize_nav_frame(frame, asset)
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if attempt < retries:
                time.sleep(retry_sleep * attempt)
    raise RuntimeError(f"FETCH_FUND_NAV_FAILED {asset.ts_code}: {last_error}")


def normalize_nav_frame(frame: pd.DataFrame, asset: FundAsset) -> pd.DataFrame:
    result = frame.copy()
    for column in FUND_NAV_FIELDS:
        if column not in result.columns:
            result[column] = pd.NA
    result = result.loc[:, FUND_NAV_FIELDS]
    result["ts_code"] = asset.ts_code
    result["symbol"] = asset.ts_code.split(".", maxsplit=1)[0]
    result["fund_name"] = asset.fund_name
    result["instrument_type"] = asset.instrument_type
    result["asset_bucket"] = asset.asset_bucket
    for column in ["ann_date", "nav_date"]:
        result[column] = pd.to_datetime(result[column], format="%Y%m%d", errors="coerce")
    numeric_columns = [
        "unit_nav",
        "accum_nav",
        "accum_div",
        "net_asset",
        "total_netasset",
        "adj_nav",
    ]
    for column in numeric_columns:
        result[column] = pd.to_numeric(result[column], errors="coerce")
    result = result.dropna(subset=["nav_date"]).drop_duplicates(["ts_code", "nav_date"], keep="last")
    result = result.sort_values(["ts_code", "nav_date"]).reset_index(drop=True)
    result["adj_nav_return"] = result.groupby("ts_code", sort=False)["adj_nav"].pct_change()
    result["unit_nav_return"] = result.groupby("ts_code", sort=False)["unit_nav"].pct_change()
    result["accum_nav_return"] = result.groupby("ts_code", sort=False)["accum_nav"].pct_change()
    ordered = [
        "ts_code",
        "symbol",
        "fund_name",
        "instrument_type",
        "asset_bucket",
        "ann_date",
        "nav_date",
        "unit_nav",
        "accum_nav",
        "accum_div",
        "net_asset",
        "total_netasset",
        "adj_nav",
        "adj_nav_return",
        "unit_nav_return",
        "accum_nav_return",
    ]
    return result.loc[:, ordered]


def build_summary(long_frame: pd.DataFrame, assets: Sequence[FundAsset]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for asset in assets:
        sub = long_frame.loc[long_frame["ts_code"] == asset.ts_code]
        rows.append(
            {
                "ts_code": asset.ts_code,
                "symbol": asset.ts_code.split(".", maxsplit=1)[0],
                "fund_name": asset.fund_name,
                "instrument_type": asset.instrument_type,
                "asset_bucket": asset.asset_bucket,
                "status": "ok" if not sub.empty else "failed",
                "rows": int(len(sub)),
                "first_nav_date": sub["nav_date"].min().date().isoformat() if not sub.empty else "",
                "last_nav_date": sub["nav_date"].max().date().isoformat() if not sub.empty else "",
                "missing_adj_nav_rows": int(sub["adj_nav"].isna().sum()) if not sub.empty else pd.NA,
                "missing_accum_nav_rows": int(sub["accum_nav"].isna().sum()) if not sub.empty else pd.NA,
                "missing_unit_nav_rows": int(sub["unit_nav"].isna().sum()) if not sub.empty else pd.NA,
            }
        )
    return pd.DataFrame(rows)


def write_outputs(
    long_frame: pd.DataFrame,
    summary: pd.DataFrame,
    manifest: dict[str, Any],
    *,
    asset_order: Sequence[str],
    excel_output: Path,
    long_csv_output: Path,
    summary_csv_output: Path,
    long_parquet_output: Path,
    wide_adj_nav_parquet_output: Path,
    manifest_output: Path,
) -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    long_frame.to_csv(long_csv_output, index=False)
    summary.to_csv(summary_csv_output, index=False)
    long_frame.to_parquet(long_parquet_output, index=False)
    columns = list(asset_order)
    wide_adj_nav = long_frame.pivot(index="nav_date", columns="ts_code", values="adj_nav").reindex(columns=columns).sort_index()
    wide_adj_nav.to_parquet(wide_adj_nav_parquet_output)
    wide_unit_nav = long_frame.pivot(index="nav_date", columns="ts_code", values="unit_nav").reindex(columns=columns).sort_index()
    wide_accum_nav = long_frame.pivot(index="nav_date", columns="ts_code", values="accum_nav").reindex(columns=columns).sort_index()
    wide_adj_nav_return = (
        long_frame.pivot(index="nav_date", columns="ts_code", values="adj_nav_return").reindex(columns=columns).sort_index()
    )
    manifest_output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    manifest_frame = pd.DataFrame(
        [
            {"key": key, "value": json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value}
            for key, value in manifest.items()
        ]
    )
    with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
        manifest_frame.to_excel(writer, sheet_name="manifest", index=False)
        summary.to_excel(writer, sheet_name="asset_summary", index=False)
        long_frame.to_excel(writer, sheet_name="fund_nav_long", index=False)
        wide_adj_nav.to_excel(writer, sheet_name="wide_adj_nav")
        wide_adj_nav_return.to_excel(writer, sheet_name="wide_adj_nav_return")
        wide_unit_nav.to_excel(writer, sheet_name="wide_unit_nav")
        wide_accum_nav.to_excel(writer, sheet_name="wide_accum_nav")
    format_excel(excel_output)


def format_excel(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=False)
        for column_index in range(1, sheet.max_column + 1):
            max_len = 0
            for row_index in range(1, min(sheet.max_row, 200) + 1):
                value = sheet.cell(row=row_index, column=column_index).value
                max_len = max(max_len, len(str(value)) if value is not None else 0)
            header = sheet.cell(row=1, column=column_index).value
            width = min(max(max_len + 2, 10), 34)
            if header in {"fund_name", "value"}:
                width = 52
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        sheet.row_dimensions[1].height = 28
    workbook.save(path)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Download Core-13 ETF/LOF fund_nav data from TuShare.")
    parser.add_argument("--universe", default=str(ROOT / "configs" / "data" / "core13_universe.yaml"))
    parser.add_argument("--start-date", default="20000101")
    parser.add_argument("--end-date", default=pd.Timestamp.now(tz="Asia/Shanghai").strftime("%Y%m%d"))
    parser.add_argument("--market", default="E", choices=["E", "O"], help="TuShare fund_nav market: E=场内, O=场外.")
    parser.add_argument(
        "--token-env",
        action="append",
        default=None,
        help="Environment variable name containing TuShare token. Can be repeated.",
    )
    parser.add_argument("--excel-output", default=str(DEFAULT_EXCEL_OUTPUT))
    parser.add_argument("--long-csv-output", default=str(DEFAULT_LONG_CSV_OUTPUT))
    parser.add_argument("--summary-csv-output", default=str(DEFAULT_SUMMARY_CSV_OUTPUT))
    parser.add_argument("--long-parquet-output", default=str(DEFAULT_LONG_PARQUET_OUTPUT))
    parser.add_argument("--wide-adj-nav-parquet-output", default=str(DEFAULT_WIDE_ADJ_NAV_PARQUET_OUTPUT))
    parser.add_argument("--manifest-output", default=str(DEFAULT_MANIFEST_OUTPUT))
    parser.add_argument("--retries", type=int, default=3)
    parser.add_argument("--retry-sleep", type=float, default=1.0)
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> None:
    args = parse_args(argv)
    universe_path = Path(args.universe)
    assets = load_universe(universe_path) if universe_path.exists() else CORE13_ASSETS
    token_env_names = args.token_env or ["TUSHARE_TOKEN", "TS_TOKEN"]
    token, token_env_name = resolve_token(token_env_names)
    pro = get_tushare_client(token)
    frames: list[pd.DataFrame] = []
    errors: list[dict[str, str]] = []
    for index, asset in enumerate(assets, start=1):
        print(f"[{index:02d}/{len(assets)}] download {asset.ts_code} {asset.fund_name}", flush=True)
        try:
            frame = fetch_asset_nav(
                pro,
                asset,
                market=args.market,
                start_date=args.start_date,
                end_date=args.end_date,
                retries=args.retries,
                retry_sleep=args.retry_sleep,
            )
            frames.append(frame)
            print(
                f"    rows={len(frame)} {frame['nav_date'].min().date()} -> {frame['nav_date'].max().date()}",
                flush=True,
            )
        except Exception as exc:  # noqa: BLE001
            print(f"    ERROR {exc}", flush=True)
            errors.append({"ts_code": asset.ts_code, "fund_name": asset.fund_name, "error": str(exc)})
    if not frames:
        raise RuntimeError("NO_TUSHARE_FUND_NAV_DATA_DOWNLOADED")

    long_frame = pd.concat(frames, ignore_index=True, sort=False)
    long_frame = long_frame.sort_values(["ts_code", "nav_date"]).reset_index(drop=True)
    summary = build_summary(long_frame, assets)
    complete_adj_nav = long_frame.pivot(index="nav_date", columns="ts_code", values="adj_nav").dropna()
    manifest = {
        "generated_at": pd.Timestamp.now(tz="Asia/Shanghai").isoformat(),
        "source": "TuShare pro.fund_nav",
        "market": args.market,
        "start_date_requested": args.start_date,
        "end_date_requested": args.end_date,
        "fields": FUND_NAV_FIELDS,
        "token_env_name": token_env_name,
        "universe_path": str(universe_path),
        "canonical_asset_order": [asset.ts_code for asset in assets],
        "asset_count_requested": len(assets),
        "asset_count_ok": int((summary["status"] == "ok").sum()),
        "long_rows": int(len(long_frame)),
        "panel_start": long_frame["nav_date"].min().date().isoformat(),
        "panel_end": long_frame["nav_date"].max().date().isoformat(),
        "complete_adj_nav_start": complete_adj_nav.index.min().date().isoformat() if not complete_adj_nav.empty else "",
        "complete_adj_nav_end": complete_adj_nav.index.max().date().isoformat() if not complete_adj_nav.empty else "",
        "complete_adj_nav_rows": int(len(complete_adj_nav)),
        "errors": errors,
        "notes": [
            "adj_nav is TuShare fund_nav 复权单位净值.",
            "No token value is written to this manifest or output files.",
        ],
    }
    write_outputs(
        long_frame,
        summary,
        manifest,
        asset_order=[asset.ts_code for asset in assets],
        excel_output=Path(args.excel_output),
        long_csv_output=Path(args.long_csv_output),
        summary_csv_output=Path(args.summary_csv_output),
        long_parquet_output=Path(args.long_parquet_output),
        wide_adj_nav_parquet_output=Path(args.wide_adj_nav_parquet_output),
        manifest_output=Path(args.manifest_output),
    )
    print("[done]")
    print(f"asset_count_ok={manifest['asset_count_ok']} of {manifest['asset_count_requested']}")
    print(f"long_rows={manifest['long_rows']}")
    print(f"panel_date_range={manifest['panel_start']} -> {manifest['panel_end']}")
    print(f"complete_adj_nav_range={manifest['complete_adj_nav_start']} -> {manifest['complete_adj_nav_end']}")
    print(f"excel_output={Path(args.excel_output)}")
    print(f"long_csv_output={Path(args.long_csv_output)}")
    print(f"long_parquet_output={Path(args.long_parquet_output)}")


if __name__ == "__main__":
    main()
